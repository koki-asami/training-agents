"""Load scenarios from training-scenario-generator output (Excel/JSON)."""

import json
import re
from pathlib import Path

import structlog

from src.models.enums import (
    DEPT_KEYWORDS,
    SOURCE_TO_AGENT,
    SOURCE_TO_RESPONSIBLE_DEPT,
    AgentRole,
    DifficultyLevel,
)
from src.models.scenario import ScenarioConfig, ScenarioEvent, TrainingLevelInfo

logger = structlog.get_logger()

# Response window adjustments by difficulty
RESPONSE_WINDOW_MULTIPLIERS = {
    DifficultyLevel.BEGINNER: 2.0,
    DifficultyLevel.INTERMEDIATE: 1.0,
    DifficultyLevel.ADVANCED: 0.75,
}

# Column name aliases: real Excel column name -> canonical field name
# Supports both the original format and the real scenario format
COLUMN_ALIASES = {
    # Event ID
    "状況付与番号": "event_id",
    "付与番号": "event_id",
    "event_id": "event_id",
    # Title
    "付与内容": "title",
    "title": "title",
    # Time
    "時刻": "scheduled_time",
    "時間": "scheduled_time",
    "scheduled_time": "scheduled_time",
    # Date
    "日付": "date",
    "date": "date",
    # Source
    "情報源": "source",
    "source": "source",
    # Content
    "内容_管理用詳細": "content_admin",
    "content_admin": "content_admin",
    "内容_訓練者向け": "content_trainee",
    "content_trainee": "content_trainee",
    # Training info
    "狙い": "training_objective",
    "training_objective": "training_objective",
    "訓練の効果": "training_effect",
    "training_effect": "training_effect",
    "期待される対応行動": "expected_actions",
    "expected_actions": "expected_actions",
    "想定される課題": "expected_issues",
    "expected_issues": "expected_issues",
    # Terrain and hazard
    "地形情報や想定される被害の特徴": "terrain_info",
    "terrain_info": "terrain_info",
    "水位状況": "water_level_status",
    "water_level_status": "water_level_status",
    "想定される二次災害のリスク": "secondary_disaster_risks",
    "secondary_disaster_risks": "secondary_disaster_risks",
    # Weather/river context
    "気象情報": "weather_info",
    "weather_info": "weather_info",
    "河川情報": "river_info",
    "river_info": "river_info",
}


def _resolve_target_agent(source: str) -> AgentRole:
    """Map 情報源 to the agent that originates this event."""
    for key, role in SOURCE_TO_AGENT.items():
        if key in source:
            return role
    return AgentRole.GENERAL_AFFAIRS


def _resolve_responsible_department(source: str, expected_actions: str) -> str:
    """Determine the responsible department from source and expected_actions.

    Priority:
    1. Keywords in expected_actions (most specific)
    2. Default mapping from information source
    """
    # Check expected_actions for department keywords
    if expected_actions:
        for keyword, dept in DEPT_KEYWORDS.items():
            if keyword in expected_actions:
                return dept

    # Fall back to source-based mapping
    for key, dept in SOURCE_TO_RESPONSIBLE_DEPT.items():
        if key in source:
            return dept

    return "総務部"


def _safe_str(value) -> str:
    """Convert a cell value to string, handling None and newlines."""
    if value is None:
        return ""
    return str(value).strip()


def _normalize_time(time_str: str) -> str:
    """Normalize time string to HH:MM format.

    Handles formats like '13:38', '8:00', datetime objects, etc.
    """
    if not time_str:
        return "00:00"

    # If it's already HH:MM
    match = re.match(r"^(\d{1,2}):(\d{2})", str(time_str))
    if match:
        h, m = int(match.group(1)), int(match.group(2))
        return f"{h:02d}:{m:02d}"

    return str(time_str)[:5]


def _map_row_to_fields(row_data: dict[str, object]) -> dict[str, str]:
    """Map a row dict (with original column names) to canonical field names."""
    mapped = {}
    for col_name, value in row_data.items():
        if col_name is None:
            continue
        canonical = COLUMN_ALIASES.get(col_name)
        if canonical:
            mapped[canonical] = _safe_str(value)
    return mapped


def _build_event(
    fields: dict[str, str], response_multiplier: float, row_index: int = 0
) -> ScenarioEvent | None:
    """Build a ScenarioEvent from mapped fields.

    Returns None only if there's no meaningful content (no title AND no content).
    Auto-generates event_id and scheduled_time if missing.
    """
    event_id = fields.get("event_id", "")
    title = fields.get("title", "")
    content_trainee = fields.get("content_trainee", "")
    content_admin = fields.get("content_admin", "")

    # Skip rows with no meaningful content at all
    if not title and not content_trainee and not content_admin:
        return None

    # Auto-generate event_id if missing
    if not event_id:
        event_id = str(row_index)

    event = ScenarioEvent(
        event_id=event_id,
        title=fields.get("title", ""),
        date=fields.get("date", ""),
        scheduled_time=_normalize_time(fields.get("scheduled_time", "00:00")),
        source=fields.get("source", ""),
        content_admin=fields.get("content_admin", ""),
        content_trainee=fields.get("content_trainee", ""),
        training_objective=fields.get("training_objective", ""),
        training_effect=fields.get("training_effect", ""),
        expected_actions=fields.get("expected_actions", ""),
        expected_issues=fields.get("expected_issues", ""),
        terrain_info=fields.get("terrain_info", ""),
        water_level_status=fields.get("water_level_status", ""),
        secondary_disaster_risks=fields.get("secondary_disaster_risks", ""),
        weather_info=fields.get("weather_info", ""),
        river_info=fields.get("river_info", ""),
        response_window_minutes=int(10 * response_multiplier),
    )
    event.target_agent = _resolve_target_agent(event.source)
    event.responsible_department = _resolve_responsible_department(
        event.source, event.expected_actions
    )
    return event


def _build_alert_timeline(events: list[ScenarioEvent]) -> list[dict]:
    """Extract alert timeline from weather_info/river_info fields across events."""
    timeline = []
    seen = set()
    for event in events:
        for info_field in [event.weather_info, event.river_info]:
            if info_field and info_field not in seen:
                seen.add(info_field)
                timeline.append({
                    "time": event.scheduled_time,
                    "info": info_field,
                })
    return timeline


def load_scenario_from_json(path: str | Path, difficulty: DifficultyLevel) -> ScenarioConfig:
    """Load a scenario from a JSON file."""
    path = Path(path)
    with open(path, encoding="utf-8") as f:
        data = json.load(f)

    response_multiplier = RESPONSE_WINDOW_MULTIPLIERS[difficulty]
    events = []

    for i, item in enumerate(data.get("events", data if isinstance(data, list) else [])):
        fields = _map_row_to_fields(item)
        event = _build_event(fields, response_multiplier, row_index=i + 1)
        if event:
            events.append(event)

    events.sort(key=lambda e: e.scheduled_time)

    config = ScenarioConfig(
        municipality=data.get("municipality", path.stem),
        difficulty=difficulty,
        events=events,
        alert_timeline=data.get("alert_timeline", _build_alert_timeline(events)),
    )

    if events:
        config.sim_start_time = events[0].scheduled_time
        config.sim_end_time = events[-1].scheduled_time

    logger.info("scenario_loaded", path=str(path), event_count=len(events))
    return config


def load_scenario_from_excel(
    path: str | Path, difficulty: DifficultyLevel
) -> ScenarioConfig:
    """Load a scenario from an Excel file (training-scenario-generator output).

    Supports the real scenario format with:
    - Sheet 'シナリオ': main events (columns B-Q)
    - Sheet '訓練レベル情報': training level metadata
    - Sheet '地理情報': GeoJSON geographic data
    - Sheet '参照シナリオ': reference scenarios
    """
    import openpyxl

    path = Path(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # --- Load シナリオ sheet ---
    scenario_sheet = None
    for name in ["シナリオ", "Sheet1"]:
        if name in wb.sheetnames:
            scenario_sheet = wb[name]
            break
    if scenario_sheet is None:
        scenario_sheet = wb.active

    # Read headers from first row
    headers = [cell.value for cell in next(scenario_sheet.iter_rows(min_row=1, max_row=1))]

    response_multiplier = RESPONSE_WINDOW_MULTIPLIERS[difficulty]
    events = []

    # Track last known date/weather/river for rows that inherit from above
    last_date = ""
    last_weather = ""
    last_river = ""

    row_index = 0
    for row in scenario_sheet.iter_rows(min_row=2, values_only=True):
        row_index += 1
        row_data = dict(zip(headers, row))
        fields = _map_row_to_fields(row_data)

        # Inherit date/weather/river from previous row if empty (merged cells)
        if fields.get("date"):
            last_date = fields["date"]
        else:
            fields["date"] = last_date

        if fields.get("weather_info"):
            last_weather = fields["weather_info"]
        else:
            fields["weather_info"] = last_weather

        if fields.get("river_info"):
            last_river = fields["river_info"]
        else:
            fields["river_info"] = last_river

        event = _build_event(fields, response_multiplier, row_index=row_index)
        if event:
            events.append(event)

    # If no events have a real scheduled_time, auto-generate at 2-minute intervals
    has_times = any(e.scheduled_time != "00:00" for e in events)
    if not has_times and events:
        base_h, base_m = 9, 0  # default start at 09:00
        for i, event in enumerate(events):
            total_min = base_h * 60 + base_m + i * 2
            event.scheduled_time = f"{total_min // 60:02d}:{total_min % 60:02d}"
        logger.info("auto_generated_times", event_count=len(events), start="09:00", interval_min=2)

    events.sort(key=lambda e: e.scheduled_time)

    # --- Extract municipality from filename ---
    # Pattern: 上益城郡嘉島町_体制構築_シナリオ_20260123_172311.xlsx
    stem = path.stem
    municipality = stem.split("_")[0] if "_" in stem else stem

    # --- Extract training level from filename or sheet ---
    training_level = ""
    parts = stem.split("_")
    if len(parts) >= 2:
        training_level = parts[1]  # e.g., "体制構築"

    # --- Load 訓練レベル情報 sheet ---
    training_level_info = None
    if "訓練レベル情報" in wb.sheetnames:
        ws_level = wb["訓練レベル情報"]
        level_data = {}
        for row in ws_level.iter_rows(values_only=True):
            if row[0]:
                level_data[str(row[0])] = _safe_str(row[1]) if len(row) > 1 else ""

        training_level_info = TrainingLevelInfo(
            training_level=level_data.get("訓練レベル", training_level),
            objective=level_data.get("狙い", ""),
            response_guidelines=level_data.get("想定のレベル及び応答要領", ""),
            event_count=int(level_data.get("状況付与数", "0") or "0"),
            related_agencies=level_data.get("関係機関", ""),
        )
        if training_level_info.training_level:
            training_level = training_level_info.training_level

    # --- Load 地理情報 sheet ---
    geojson_data = ""
    if "地理情報" in wb.sheetnames:
        ws_geo = wb["地理情報"]
        geo_parts = []
        for row in ws_geo.iter_rows(min_row=2, values_only=True):
            if row[0]:
                geo_parts.append(str(row[0]))
        if geo_parts:
            geojson_data = "\n".join(geo_parts)

    wb.close()

    # Build alert timeline from weather/river info changes
    alert_timeline = _build_alert_timeline(events)

    config = ScenarioConfig(
        municipality=municipality,
        training_level=training_level,
        training_level_info=training_level_info,
        difficulty=difficulty,
        events=events,
        alert_timeline=alert_timeline,
        geojson_data=geojson_data,
    )

    if events:
        config.sim_start_time = events[0].scheduled_time
        config.sim_end_time = events[-1].scheduled_time

    logger.info(
        "scenario_loaded_excel",
        path=str(path),
        municipality=municipality,
        training_level=training_level,
        event_count=len(events),
    )
    return config
